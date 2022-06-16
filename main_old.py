import datetime
from typing import Dict, Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import config
from logging_format import *
import re
import psycopg2
from multipledispatch import dispatch
import json
import math
import os


class TYPEErrors(object):

    def __init__(self, df: pd.DataFrame, name_table: str):
        super(TYPEErrors, self).__init__()
        self.df = df
        self.name_table = name_table
        self.type_error = ''
        self.dict_json = {}
        # self.data = {}
        self.data = {'Very big value': [], 'Very small value': [], 'Wrong sign': [], 'Small data': [], 'Big data': [],
                     'Error span year': [], 'Mismatch of values in two tables (new value is lower then old)': [],
                     'Dublicate rows in resassagment': []}

    def find_max_type(self, type_columns_):
        dict_type_column = dict(zip(type_columns_, [type_columns_.count(i) for i in type_columns_]))
        max_type = ''
        if type_columns_:
            max_type = max(dict_type_column, key=dict_type_column.get)
        return max_type, dict_type_column

    def write_to_dict(self):
        # self.dict_json = 0
        with open('/'.join((PATH_TO_LOG, 'log.txt')), 'a+') as outfile:
            json.dump(self.data, outfile)
            outfile.write(', \n')

    def func_all_int_float_array(self, column1, type_columns) -> list:
        return [column1[i] for i, value in enumerate(column1) if
                (type_columns[i] == int or type_columns[i] == float or type_columns[i] == np.float32
                 or type_columns[i] == np.float64 or type_columns[i] == np.int64 or type_columns[i] == np.int32)]

    def sign_warrning(self, column_index, begin_table, column1, type_columns, name_table, df, PK) -> list:
        all_int_float_array = self.func_all_int_float_array(column1, type_columns)
        minus_int_float_array = [i for i in all_int_float_array if i < 0]
        error_sign = []
        if len(all_int_float_array) > len(minus_int_float_array) >= 1 and len(minus_int_float_array) < len(
                column1) // 10:
            # print(f'dict_type_column = {dict_type_column}')
            for row, elem in enumerate(column1):
                if elem in minus_int_float_array:
                    self.print_name_table_row_PK(df, column_index, name_table,
                                                 df.values[row + begin_table - 1][column_index],
                                                 row, True, PK, 'Wrong sign')

                    error_sign.append([column_index, row])
        return error_sign

    def big_small_value_warrning(self, column_index, begin_table, column1, type_columns, error_sign, name_table, df,
                                 PK):
        all_int_float_array = self.func_all_int_float_array(column1, type_columns)
        if len(all_int_float_array) != column1.shape[0]:
            print('Nane in column1')

        error_sign_values = []
        if error_sign:
            minus_values = [error_sign[j][1] for j in range(len(error_sign)) if column_index == error_sign[j][0]]
            error_sign_values = [column1[i] for i in minus_values]
        try:
            uniq_resource = np.unique(df['id_resource'])
            for id_resource in uniq_resource:
                index = [i for i, value in enumerate(df['id_resource']) if value == id_resource]

                quant = np.quantile(column1[index], 0.9)
                quant_05 = np.quantile(column1[index], 0.5)
                k = quant * len(column1[index]) // (10 ** (len(str(len(column1[index]))) - 1))

                if k == 0:
                    k = quant
                print('k = ', k)
                for row, elem in enumerate(column1[index]):
                    if elem in all_int_float_array and elem not in error_sign_values:

                        if elem > k:
                            self.print_name_table_row_PK(df, column_index, name_table, elem, index[row], True, PK,
                                                         'Very big value')
                            logger.warning(
                                f'warning big value {elem} in cell : column = {column_index} row = {index[row]}')

                        if elem < quant // len(column1[index]) // 10:
                            print("elem = ", elem)
                            print("quant // len(column1[index]) // 10 = ", quant // len(column1[index]) // 10)
                            self.print_name_table_row_PK(df, column_index, name_table, elem, index[row], True, PK,
                                                         'Very small value')

                            # # print(f'elem = {elem}, row = {row}, column_ind = {column_index}')
                            logger.warning(
                                f'warning small value {elem} in cell : column = {column_index} row = {index[row]}')

        except:
            quant = np.quantile(column1, 0.9)
            np.quantile(column1, 0.5)
            print(f'QUAINTIL = {quant}')
            quant_05 = np.quantile(column1, 0.5)
            for row, elem in enumerate(column1):
                if elem in all_int_float_array and elem not in error_sign_values:

                    if elem > quant * len(column1) // 100:  # (10**(len(str(len(column1)))-1)):

                        self.print_name_table_row_PK(df, column_index, name_table, elem, row, True, PK,
                                                     'Very big value')
                        logger.warning(
                            f'warning big value {elem} in cell : column = {column_index} row = {row}')

                    if elem < quant // len(column1) // 10:
                        self.print_name_table_row_PK(df, column_index, name_table, elem, row, True, PK,
                                                     'Very small value')

                        logger.warning(
                            f'warning small value {elem} in cell : column = {column_index} row = {row}')

    def check_date(self, column1, column_index, name_table, df, PK):
        date_today = datetime.date.today()
        too_big_year = [i for i, value in enumerate(column1) if value.year > date_today.year + 10]
        too_small_year = [i for i, value in enumerate(column1) if value.year < date_today.year - 10]
        if too_big_year:
            for i in too_big_year:
                self.print_name_table_row_PK(df, column_index, name_table, column1, i, True, PK, 'Big data')
                logger.warning(
                    f'warning big year in cell : column = {column_index} row = {i}')
        if too_small_year:
            for i in too_small_year:
                self.print_name_table_row_PK(df, column_index, name_table, column1, i, True, PK, 'Small data')
                logger.warning(
                    f'warning small year in cell : column = {column_index} row = {i}')

    def conv_type_for_json(self, column1, name_table, two_pk):
        dict_value = {}
        if two_pk:
            print(f'value = {column1[0]}')
            print(f'value_pmc = {column1[1]}')
            if type(column1[0]) == int:
                dict_value = {
                    f'value': int(column1[0]),
                    f'pmc value': int(column1[1])}
            elif type(column1[0]) == float or type(column1[0]) == np.float64 or type(column1[0]) == np.float32:
                dict_value = {
                    f'value': float(column1[0]),
                    f'pmc value': float(column1[1])}
        else:
            print(f'value = {column1}')
            if type(column1) == np.float64 or type(column1) == np.float32 or type(column1) == float:
                dict_value = {f'value': float(column1)}
            elif type(column1) == np.int:
                dict_value = {f'value': int(column1)}

        return dict_value

    def print_name_table_row_PK(self, df, i, name_table, column1, j, one_table, PK, type_error):
        dict_ = {'name_table': name_table,
                 'name_column': df.columns[i]}
        dict_PK = {}
        dict_value = {}
        print(f'name column = {df.columns[i]}')
        if not one_table:
            if type(PK) == str:
                print(f'PK {PK} = {df[PK].values[j][0]}')
                dict_PK = {'PK': int(df[PK].values[j][0])}
            else:
                print(f'PK {PK[0]} = {df[PK[0]].values[j][0]}')
                print(f'PK2 {PK[1]} = {df[PK[1]].values[j][0]}')
                dict_PK = {
                    PK[0]: int(df[PK[0]].values[j][0]),
                    PK[1]: str(df[PK[1]].values[j][0])}
            print(f'{name_table} value = {column1[j][i]}')
            print(f'{name_table}_pmc value = {column1[j][i + (df.shape[1]) // 2]} \n')
            values = [column1[j][i], column1[j][i + (df.shape[1]) // 2]]
            dict_value = self.conv_type_for_json(values, name_table, True)
        else:
            if type(PK) == str:
                print(f'PK {PK} = {df[PK].values[j]}')
                dict_PK = {'PK': int(df[PK].values[j])}
            else:
                print(f'PK {PK[0]} = {df[PK[0]].values[j]}')
                print(f'PK2 {PK[1]} = {df[PK[1]].values[j]}')
                dict_PK = {
                    PK[0]: int(df[PK[0]].values[j]),
                    PK[1]: str(df[PK[1]].values[j])}
            if type(column1) == np.ndarray:
                dict_value = self.conv_type_for_json(column1[j], name_table, False)
            else:
                dict_value = self.conv_type_for_json(column1, name_table, False)
        c = {**dict_, **dict_PK, **dict_value}
        self.data[f'{type_error}'].append(c)

    def analysis_resassagnment(self, i):
        if i == 'resassignment':
            sql_dubic_resassignment = f""" select {i}.id_resassignment, {i}.id_resource, {i}.id_activity, {i}.id_project,{i}.id_wbs
             from {i} """
            df_resassignment_dublicates = pd.read_sql_query(sql_dubic_resassignment, connection)
            # df_resassignment_dublicates = pd.read_excel(PATH_READ)
            df1 = df_resassignment_dublicates[['id_resource', 'id_project', 'id_wbs', 'id_activity']]
            rm_dublicates_df = df1.drop_duplicates()
            if rm_dublicates_df.shape != df1.shape:
                print('Find_ diblicates')

                df_find_dubl = df1.drop(rm_dublicates_df.index)
                rm = df_find_dubl.drop_duplicates()
                dubl = rm[['id_project', 'id_resource', 'id_wbs', 'id_activity']].values
                # dubl = df_find_dubl[['id_project', 'id_resource', 'id_wbs', 'id_activity']].values
                for i_dubl in dubl:
                    l = list(df1.loc[(df_resassignment_dublicates[['id_project', 'id_resource', 'id_wbs',
                                                                   'id_activity']] == i_dubl).all(axis=1)].index)

                    ind_list = [df_resassignment_dublicates['id_resassignment'][i_ind] for i_ind in l]
                    #
                    print("ind_list = ", ind_list)
                    print("l",l)
                    print("len(ind_list)-1  = " ,len(ind_list)-1)
                    for i in range(0, len(ind_list), 1):
                        # print(l[i])
                        print(i)
                        self.print_name_table_row_PK(df_resassignment_dublicates, 0, 'resassignment', [], l[i],
                                                     True,'id_resassignment', 'Dublicate rows in resassagment')
                        # logger.warning(f'Dublicate rows id = {ind_list[i]} and id = {ind_list[i + 1]}')

    def analysis_two_tables(self, df, PK, name_table):
        column1 = df.values
        for j in range(df.shape[0]):
            for i in range(0, df.shape[1] // 2, 2):
                if df.columns[i] != 'update_date':
                    # if column1[j][i] != column1[j][i+1]:
                    if column1[j][i] != column1[j][i + (df.shape[1]) // 2]:
                        if type(column1[j][i]) != pd._libs.tslibs.nattype.NaTType and \
                                type(column1[j][i + (df.shape[1]) // 2]) != pd._libs.tslibs.nattype.NaTType:
                            if type(column1[j][i]) == pd._libs.tslibs.timestamps.Timestamp:
                                self.print_name_table_row_PK(df, i, name_table, column1, j, False, PK,
                                                             'Mismatch of values in two tables')
                            # ошибка! тип vakue не timestamp
                            if type(column1[j][i]) != pd._libs.tslibs.timestamps.Timestamp:
                                if not math.isnan(column1[j][i]) and not math.isnan(column1[j][i + (df.shape[1]) // 2]):
                                    if column1[j][i] > column1[j][i + (df.shape[1]) // 2]:
                                        self.print_name_table_row_PK(df, i, name_table, column1, j, False, PK,
                                                                     'Mismatch of values in two tables (new value is lower then old)')

    # @dispatch(object, str, str)
    def analis_data_df(self, df, name_table, PK):
        start_date_column = []
        begin_table = 1
        dtypes = df.dtypes
        for column_index in range(df.shape[1]):
            id_ = re.findall(r'(\w{2}_)', df.columns[column_index])
            id = re.findall(r'(\w{2})', df.columns[column_index])
            word_id = re.findall(r'(_\w{2})', df.columns[column_index])
            if id_:
                if id_[0] == 'id_':
                    continue
            if id:
                if id[0] == 'id':
                    continue
            if word_id:
                if word_id[len(word_id) - 1] == '_id':
                    continue
            if dtypes[column_index] == object or dtypes[column_index] == bool:
                continue
            # print(df.columns[column_index])
            column1 = df[df.columns[column_index]].values
            type_columns_ = [type(column1[i]) for i in range(len(column1)) if column1[i] != '']
            if type_columns_:

                type_columns_ = [type(column1[i]) for i in range(len(column1)) if
                                 type(column1[i]) != pd._libs.tslibs.nattype.NaTType or np.isnat(column1[i])]
                type_columns = [type(column1[i]) for i in range(len(column1))]
                max_type, dict_type_column = self.find_max_type(type_columns_)
                # error_type = find_type_error(max_type, type_columns, df, begin_table,
                #                              column_index, name_table)
                if max_type == int or max_type == float or max_type == np.float64 or max_type == np.float32 \
                        or max_type == np.int64 or max_type == np.int32:

                    error_sign = self.sign_warrning(column_index, begin_table, column1, type_columns, name_table, df,
                                                    PK)

                    std_ = np.std(self.func_all_int_float_array(column1, type_columns))
                    if std_ > 100:
                        self.big_small_value_warrning(column_index, begin_table, column1, type_columns, error_sign,
                                                      name_table, df, PK)

                if max_type == pd._libs.tslibs.timestamps.Timestamp or max_type == np.datetime64:
                    self.check_date([pd.to_datetime(column1[i]).date() for i in range(len(column1))], column_index,
                                    name_table, df, PK)
                    start_ = re.findall(r'(StartDate)', df.columns[column_index])
                    finish_ = re.findall(r'(FinishDate)', df.columns[column_index])
                    if start_:
                        start_date_column = column1
                    if finish_:
                        finish_date_column = column1
                        if start_date_column.all() > finish_date_column.all():
                            error_date_interval = [i for i, value in enumerate(start_date_column)
                                                   if value > finish_date_column[i]]
                            for i in error_date_interval:
                                self.print_name_table_row_PK(df, column_index, name_table,
                                                             df.values[i],
                                                             i, column_index, 'Error span year')
                                logger.warning(
                                    f'warning check span year in cell : column = {column_index} row = {i}')


dict_same_types = {int: [float], float: [int]}
ALL = True
BD = True
PATH_TO_LOG = 'log/'
# PATH_READ = 'D:\\work2\\analysis_exel_data\\analys_test.xlsx'
DICT_TABLE_PK = {'resource': 'id_resource',
                 'activity': 'id_activity',
                 'udf_code_resassignment': 'id_resassignment',
                 'project': 'id_project',
                 'udf_code_activity': 'id_activity',
                 'udf_code_resource': 'id_resource',
                 'resassignmentspred': ['id_resassignment', 'dt'],
                 'resassignment': 'id_resassignment',
                 'udf_code_project': 'id_project',
                 'dictionary_project_code': 'id',
                 'dictionary_activity_code': 'id',
                 'dictionary_resource_code': 'id',
                 'actvrel': 'id_actvrel'}
def chek_dictionary_activity_code_and_activity():
    sql = '''
            SELECT
              activity.id_activity,
              activity.activity_code,
              activity.activity_name,
              dictionary_activity_code."Dictionary_Name"
            FROM
              public.activity activity
              INNER JOIN public.dictionary_activity_code dictionary_activity_code ON dictionary_activity_code.code = activity.activity_code
            WHERE
              dictionary_activity_code."Dictionary_Name" IN ('Виды работ', 'Конструктив для отчета');
            '''
    quary_ = pd.read_sql_query(sql, connection)
    if quary_.shape[0] == 0:
        logger.warning("Ни на одну работу не назначены коды!")
        with open('/'.join((PATH_TO_LOG, 'log.txt')), 'a+') as outfile:
            outfile.write('Warning (No codes are assigned to any table!) \n')


def chek_all_tables(cursor):
    cursor.execute("""SELECT relname FROM pg_class WHERE relkind='r'
                                  AND relname !~ '^(pg_|sql_)';""")  # "rel" is short for relation.
    tables_ = [i[0] for i in cursor.fetchall()]
    tables_ = ['resassignment', 'activity']

    for i in tables_:
        if i != 'ISR' and i != 'ISR_del' and i != 'ISR_pmc' and i != 'alembic_version':
            if '_pmc' not in i and '_del' not in i:
                print("Table = ", i)
                sql_one_table = f""" select * from {i} """
                d_all = pd.read_sql_query(sql_one_table, connection)
                PK = DICT_TABLE_PK.get(i)
                type_error = TYPE_Errors(d_all, i)
                with open('/'.join((PATH_TO_LOG, 'log.txt')), 'a+') as outfile:
                    outfile.write('{"Table": [ \n')
                    outfile.write(f'"{i}", ')
                type_error.analis_data_df(d_all, i, PK)
                if i == 'resassignment':
                    type_error.analysis_resassagnment(i)

                try:
                    i_pmc = i + '_pmc'
                    # PK = DICT_TABLE_PK.get(i)
                    if type(PK) == str:  # < 2:
                        sql = f"""
                                        select * from {i}, {i}_pmc
                                        where {i}.{PK}={i}_pmc.{PK}
                                        """
                    else:
                        sql = f"""
                                        select * from {i}, {i}_pmc
                                        where {i}.{PK[0]}={i}_pmc.{PK[0]} and {i}.{PK[1]}={i}_pmc.{PK[1]}
                                         """
                    print('ANALYSIS OF TWO TABLE ORIG AND ORIG_PMC')
                    d_two_table = pd.read_sql_query(sql, connection)
                    type_error.analysis_two_tables(d_two_table, PK, i)


                except Exception as exp:
                    print(f'Error {exp}')
                    pass
                type_error.write_to_dict()
                with open('/'.join((PATH_TO_LOG, 'log.txt')), 'a+') as outfile:
                        outfile.write(']}, \n')


try:
    # if not BD:
    #     rb = load_workbook(PATH_READ, read_only=False)  # , formatting_info=True)
    #     sheet = rb['Лист1']
    #     rb.active = 1
    #     df = pd.read_excel(PATH_READ)
    #     type_error = TYPE_Errors(df, 'resassignment')
    #     type_error.analysis_resassagnment('resassignment')
    #     type_error.write_to_dict()
    # elif BD:
        # connection to bd
    connection = psycopg2.connect(host=config.host, user=config.user, password=config.password, database=config.bd_name)

    chek_dictionary_activity_code_and_activity()

    cursor = connection.cursor()
    if not os.path.exists(PATH_TO_LOG):
        os.makedirs(PATH_TO_LOG)
    with open('/'.join((PATH_TO_LOG, 'log.txt')), 'a+') as outfile:
        outfile.write('[ \n')
    chek_all_tables(cursor)
    with open('/'.join((PATH_TO_LOG, 'log.txt')), 'a+') as outfile:
        outfile.write('] \n')

except Exception as exp:
    print(f'Error {exp}')
finally:
    if connection:
        connection.close()
        print('connection is clossed')
